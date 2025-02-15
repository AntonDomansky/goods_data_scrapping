import os
import datetime
import json
from openpyxl import Workbook, load_workbook
from settings import files_directory, logs_directory


def is_path_exists(pathname):
    is_exist = os.path.exists(pathname)
    return is_exist


def create_folder(pathname):
    if is_path_exists(pathname):
        save_to_logs(f"The '{pathname}' folder already exists")
    else:
        os.mkdir(pathname)
        save_to_logs(f"The '{pathname}' folder has been created")


def save_to_logs(message, is_indent=False):
    indent = ''
    if is_indent:
        indent = '\n'
    if not is_path_exists(logs_directory):
        create_folder(logs_directory)
    now = datetime.datetime.now()
    file_name = f'{now.strftime("%Y%m%d")}_log.txt'
    with open(f"{logs_directory}/{file_name}", 'a', encoding='utf-8') as file:
        file.write(f"{indent}{now.strftime('%Y-%m-%d %H:%M:%S')}: {message}\n")
    print(f"{indent}{datetime.datetime.now().strftime('%H:%M:%S')}: {message}")


def save_to_json(data: list):
    save_to_logs(f"Saving to json-file...", True)
    create_folder(files_directory)
    try:
        with open(f"{files_directory}/parsed_data.json", 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
        save_to_logs(
            f"Data has been succesfully saved to 'parsed_data.json'")
        return True
    except Exception as e:
        save_to_logs(f"ERROR during data to json-file saving:\n{e}")
        return False


def json_to_list(json_file_name):
    save_to_logs(f"Reading json to list", True)
    try:
        with open(f"{files_directory}/{json_file_name}", encoding='utf-8') as file:
            output_list = json.load(file)
            save_to_logs(
                f"Done! Ads qty in list: {len(output_list)}")
            return output_list
    except Exception as e:
        save_to_logs(f"ERROR during json-file to list converting:\n{e}")
        return False


def write_to_xlsx(json_file_name):
    data_list = json_to_list(json_file_name)
    if data_list:
        try:
            if not is_path_exists(f"{files_directory}\parsed_data.xlsx"):
                save_to_logs(
                    f"Creating new 'parsed_data.xlsx' file and writing data", True)
                wb = Workbook()
                worksheet = wb.active
                worksheet.title = f"{datetime.datetime.now().strftime('%Y%m%d')}"
                table_header = ['№', 'Title', 'Link',
                                'Price', 'Description', 'Img Link']
                worksheet.append(table_header)
                worksheet.column_dimensions["A"].width = 5
                worksheet.column_dimensions["B"].width = 25
                worksheet.column_dimensions["C"].width = 55
                worksheet.column_dimensions["E"].width = 50
                worksheet.column_dimensions["F"].width = 45
                last_row_value = 0
            else:
                wb = load_workbook(f"{files_directory}\parsed_data.xlsx")
                worksheet = wb.active
                last_row = worksheet.max_row
                last_row_value = int(worksheet[f"A{last_row}"].value)
                save_to_logs(
                    f"Excel file is already exist, strings in file: {last_row}. Appending new data")

            for i, ad in enumerate(data_list):
                number = last_row_value + i + 1
                title = ad.get('product_title', None)
                link = ad.get('product_link', None)
                price = ad.get('product_price', None)
                description = ad.get('product_description', None)
                img = ad.get('product_img', None)
                worksheet.append(
                    [number, title, link, price, description, img])

            wb.save(f"{files_directory}\parsed_data.xlsx")

            save_to_logs(f"DONE!")
        except Exception as e:
            save_to_logs(f"ERROR during xlsx writing:\n{e}")
    else:
        save_to_logs(f"No data to work with, check the json-file")
        return False


if __name__ == '__main__':
    write_to_xlsx('parsed_data.json')
    pass
